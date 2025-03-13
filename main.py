from PyPDF2 import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from io import BytesIO

from openpyxl import load_workbook


def convert_to_thai_number(number_str):
    """
    Converts a standard number string to a Thai numeral string.

    Parameters:
        number_str (str): A string containing standard Arabic numerals (e.g., "123456").

    Returns:
        str: A string containing Thai numerals (e.g., "๑๒๓๔๕๖").
    """
    # Mapping of Arabic numerals to Thai numerals
    arabic_to_thai = {
        '0': '๐',
        '1': '๑',
        '2': '๒',
        '3': '๓',
        '4': '๔',
        '5': '๕',
        '6': '๖',
        '7': '๗',
        '8': '๘',
        '9': '๙'
    }

    # Convert each character in the input string using the mapping
    thai_number_str = ''.join(arabic_to_thai.get(char, char) for char in number_str)

    return thai_number_str



def get_excel_data(file_path):

    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook["Sheet1"]

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data


file_path = "name_list.xlsx"
list_data = get_excel_data(file_path)

print(list_data)
print(list_data[0][1])
# import pdb; pdb.set_trace()

remove_background = True

FONT_NAME = "DSN-LaiThai"
FONT_SIZE = 20 

font_path = "DSN-LaiThai.ttf"  # Path to your Thai font file
pdfmetrics.registerFont(TTFont(FONT_NAME, font_path))

existing_pdf_path = "examples.pdf"  # Path to your existing PDF
output_pdf_path = "output_with_thai_text.pdf"

existing_pdf = PdfReader(existing_pdf_path)
output_pdf = PdfWriter()

# read pdf file in first page
page = existing_pdf.pages[0]

page.rotate(90)

page_width = float(page.mediabox.upper_right[0])
page_height = float(page.mediabox.upper_right[1])


print(page_width)
print(page_height)


for data in list_data:
    # create canvas for editing PDF file 
    packet = BytesIO()  # In-memory buffer to hold the overlay PDF
    can = canvas.Canvas(packet, pagesize=(page_width, page_height))

    # Set the font to the registered Thai font
    can.setFont(FONT_NAME, FONT_SIZE)  # Font size 16

    can.translate(page_width, 0)
    can.rotate(90)

    can.saveState()

    # insert running number

    

    running_number = convert_to_thai_number(data[0])
    can.drawString(500, 368, running_number)

    # insert name of student
    name = f"{data[1]}{data[2]} {data[3]}" 
    text_width = can.stringWidth(name, FONT_NAME, FONT_SIZE)
    page_center = (page_height - text_width) / 2
    can.drawString(page_center, 265, name)  # Position (x=50, y=750)

    # insert birth date
    birthNum = convert_to_thai_number(str(data[4]))
    birthMonth = data[5] 
    birthYear = convert_to_thai_number(str(data[6]))

    can.drawString(196, 232, birthNum)
    can.drawString(269, 232, birthMonth)
    can.drawString(405, 232, birthYear)

    # insert school name
    school_name = "โรงเรียนบ้านโพนแท่น"
    can.drawString(150, 178, school_name)

    # insert province name
    province_name = "ร้อยเอ็ด"
    can.drawString(155, 153, province_name)


    # insert office
    office_name = "สพป. ร้อยเอ็ด เขต ๒"
    can.drawString(310, 153, office_name)

    # insert graduated date
    graduated_date = "๓๑"
    graduated_month = "มีนาคม"
    graduated_year = "๒๕๖๘"

    can.drawString(195, 126, graduated_date)
    can.drawString(285, 126, graduated_month)
    can.drawString(400, 126, graduated_year)

    # insert signature end
    dotted_line = 90*"."
    dotted_width = can.stringWidth(dotted_line, FONT_NAME, FONT_SIZE)
    dotted_page_center = (page_height - dotted_width) / 2
    can.drawString(dotted_page_center, 60, dotted_line)

    # insert name head teacher
    head_teacher_name = "(นางสาวอำพร วรวงษ์)"
    head_teacher_width = can.stringWidth(head_teacher_name, FONT_NAME, FONT_SIZE)
    head_page_center = (page_height - head_teacher_width) / 2
    can.drawString(head_page_center, 37, head_teacher_name)


    # insert position
    position_name = "รักษาการในตำแหน่งผู้อำนวยการโรงเรียนบ้านโพนแท่น"
    position_name_width = can.stringWidth(position_name, FONT_NAME, FONT_SIZE)
    position_page_center = (page_height - position_name_width) / 2
    can.drawString(position_page_center, 13, position_name)


    can.restoreState()

    # Save the overlay PDF to the in-memory buffer
    can.save()

    # Move the buffer's pointer to the beginning so it can be read
    packet.seek(0)

    if not remove_background:

        existing_pdf = PdfReader(existing_pdf_path)
        
        # read pdf file in first page
        page = existing_pdf.pages[0]

        page.rotate(90)

        # Merge the overlay PDF (with Thai text) onto the existing PDF page
        overlay_pdf = PdfReader(packet)
        page.merge_page(overlay_pdf.pages[0])        

        # Add the modified page to the output PDF
        output_pdf.add_page(page)
    else:
       overlay_pdf = PdfReader(packet)
       page = overlay_pdf.pages[0]
       page.rotate(90)
       output_pdf.add_page(page) 


# Write the final PDF to disk
with open(output_pdf_path, "wb") as output_file:
    output_pdf.write(output_file)

print(f"Thai text has been added to the PDF. Output saved as: {output_pdf_path}")